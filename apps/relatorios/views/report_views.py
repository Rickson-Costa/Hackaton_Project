from django.views.generic import TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.contrib import messages
from datetime import datetime, timedelta
from apps.projetos.models.projeto import Projeto
from apps.contratos.models.contrato import Contrato
from apps.contratos.models.item_contrato import ItemContrato

# Para geração de arquivos
import io
import csv
import json

# PDF generation
try:
    from fpdf import FPDF
    PDF_AVAILABLE = True
    PDF_LIBRARY = 'fpdf'
except ImportError:
    try:
        from weasyprint import HTML, CSS
        from django.template.loader import render_to_string
        PDF_AVAILABLE = True
        PDF_LIBRARY = 'weasyprint'
    except ImportError:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            PDF_AVAILABLE = True
            PDF_LIBRARY = 'reportlab'
        except ImportError:
            PDF_AVAILABLE = False
            PDF_LIBRARY = None

# Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class RelatorioListView(LoginRequiredMixin, TemplateView):
    template_name = 'relatorios/relatorio_list.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Estatísticas rápidas para a dashboard
        context.update({
            'total_projetos': Projeto.objects.count(),
            'total_contratos': Contrato.objects.count(),
            'total_pagamentos': ItemContrato.objects.filter(valor_pago__gt=0).count(),
            'valor_total_recebido': ItemContrato.objects.aggregate(total=Sum('valor_pago'))['total'] or 0,
        })
        
        return context

class RelatorioProjetosView(LoginRequiredMixin, TemplateView):
    template_name = 'relatorios/relatorio_projetos.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Dados completos de projetos
        projetos = Projeto.objects.all().order_by('-data_inicio')
        
        # Resumos estatísticos
        total_projetos = projetos.count()
        projetos_ativos = projetos.filter(situacao__in=['1', '2']).count()
        projetos_concluidos = projetos.filter(situacao='6').count()
        valor_total = projetos.aggregate(total=Sum('valor'))['total'] or 0
        
        context.update({
            'projetos': projetos,
            'total_projetos': total_projetos,
            'projetos_ativos': projetos_ativos,
            'projetos_concluidos': projetos_concluidos,
            'valor_total': valor_total,
        })
        
        return context

class RelatorioProjetosGenerateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        format_type = request.GET.get('format', 'html')
        
        if format_type == 'pdf':
            return self.generate_pdf()
        elif format_type == 'excel':
            return self.generate_excel()
        else:
            return redirect('relatorios:relatorio_projetos')
    
    def generate_pdf(self):
        if not PDF_AVAILABLE:
            # Fallback para CSV se PDF não estiver disponível
            return self.generate_csv_fallback()
            
        if PDF_LIBRARY == 'fpdf':
            return self.generate_pdf_fpdf()
        else:
            # Fallback para CSV se não conseguir gerar PDF
            return self.generate_csv_fallback()
    
    def generate_pdf_fpdf(self):
        # Gerar PDF para relatório de projetos usando FPDF
        try:
            from fpdf import FPDF
            from fpdf.enums import XPos, YPos
            
            pdf = FPDF()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)
            
            # Título
            pdf.set_font('helvetica', 'B', 16)
            pdf.cell(0, 10, 'RELATÓRIO DE PROJETOS - FUNETEC', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            
            # Data de geração
            pdf.set_font('helvetica', '', 10)
            pdf.cell(0, 5, f'Gerado em: {timezone.now().strftime("%d/%m/%Y às %H:%M")}', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            pdf.ln(10)
            
            # Cabeçalho da tabela
            pdf.set_font('helvetica', 'B', 9)
            pdf.set_fill_color(8, 149, 97)  # Verde FUNETEC
            pdf.set_text_color(255, 255, 255)  # Texto branco
            
            # Larguras das colunas
            col_widths = [25, 60, 25, 25, 25, 30]
            headers = ['Código', 'Nome', 'Valor', 'Data Início', 'Data Fim', 'Situação']
            
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 8, header, border=1, align='C', fill=True)
            pdf.ln()
            
            # Dados dos projetos
            projetos = Projeto.objects.all().order_by('-data_inicio')
            
            situacao_map = {
                '1': 'Aguardando Início', '2': 'Em Andamento', '3': 'Paralisado',
                '4': 'Suspenso', '5': 'Cancelado', '6': 'Concluído'
            }
            
            pdf.set_text_color(0, 0, 0)  # Texto preto
            pdf.set_font('helvetica', '', 8)
            
            for projeto in projetos:
                # Verificar se precisa de nova página
                if pdf.get_y() > 270:
                    pdf.add_page()
                    # Repetir cabeçalho
                    pdf.set_font('helvetica', 'B', 9)
                    pdf.set_fill_color(8, 149, 97)
                    pdf.set_text_color(255, 255, 255)
                    for i, header in enumerate(headers):
                        pdf.cell(col_widths[i], 8, header, border=1, align='C', fill=True)
                    pdf.ln()
                    pdf.set_text_color(0, 0, 0)
                    pdf.set_font('helvetica', '', 8)
                
                # Dados da linha
                nome = projeto.nome[:25] + '...' if len(projeto.nome) > 25 else projeto.nome
                valor = f"R$ {projeto.valor:,.2f}" if projeto.valor else "R$ 0,00"
                data_inicio = projeto.data_inicio.strftime('%d/%m/%Y') if projeto.data_inicio else 'N/A'
                data_fim = projeto.data_encerramento.strftime('%d/%m/%Y') if projeto.data_encerramento else 'N/A'
                situacao = situacao_map.get(projeto.situacao, 'N/A')
                
                row_data = [projeto.cod_projeto, nome, valor, data_inicio, data_fim, situacao]
                
                for i, data in enumerate(row_data):
                    align = 'R' if i == 2 else 'C'  # Valor alinhado à direita
                    pdf.cell(col_widths[i], 6, str(data), border=1, align=align)
                pdf.ln()
            
            # Gerar o PDF (método correto para fpdf2)
            pdf_output = pdf.output()
            if isinstance(pdf_output, (bytearray, bytes)):
                pdf_bytes = bytes(pdf_output)
            else:
                pdf_bytes = pdf_output.encode('latin-1')
            
            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="relatorio_projetos.pdf"'
            
            return response
            
        except Exception as e:
            # Se falhar a geração de PDF, usar fallback CSV
            return self.generate_csv_fallback()
    
    def generate_csv_fallback(self):
        # Fallback CSV se PDF não funcionar
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="relatorio_projetos.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Código', 'Nome', 'Valor', 'Data Início', 'Data Encerramento', 'Situação'])
        
        projetos = Projeto.objects.all().order_by('-data_inicio')
        
        situacao_map = {
            '1': 'Aguardando Início', '2': 'Em Andamento', '3': 'Paralisado',
            '4': 'Suspenso', '5': 'Cancelado', '6': 'Concluído'
        }
        
        for projeto in projetos:
            writer.writerow([
                projeto.cod_projeto,
                projeto.nome,
                f"R$ {projeto.valor:,.2f}" if projeto.valor else "R$ 0,00",
                projeto.data_inicio.strftime('%d/%m/%Y') if projeto.data_inicio else 'N/A',
                projeto.data_encerramento.strftime('%d/%m/%Y') if projeto.data_encerramento else 'N/A',
                situacao_map.get(projeto.situacao, 'N/A')
            ])
        
        return response
    
    def generate_excel(self):
        if not EXCEL_AVAILABLE:
            # Fallback para JSON se Excel não estiver disponível
            return self.generate_json_fallback()
            
        # Gerar Excel para relatório de projetos
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Projetos"
        
        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="089561", end_color="089561", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        # Título
        ws.merge_cells('A1:F1')
        ws['A1'] = "RELATÓRIO DE PROJETOS - FUNETEC"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Data de geração
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Gerado em: {timezone.now().strftime('%d/%m/%Y às %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Cabeçalhos das colunas
        headers = ['Código', 'Nome', 'Valor', 'Data Início', 'Data Encerramento', 'Situação']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Dados dos projetos
        projetos = Projeto.objects.all().order_by('-data_inicio')
        
        situacao_map = {
            '1': 'Aguardando Início', '2': 'Em Andamento', '3': 'Paralisado',
            '4': 'Suspenso', '5': 'Cancelado', '6': 'Concluído'
        }
        
        for row, projeto in enumerate(projetos, 5):
            ws.cell(row=row, column=1, value=projeto.cod_projeto)
            ws.cell(row=row, column=2, value=projeto.nome)
            ws.cell(row=row, column=3, value=f"R$ {projeto.valor:,.2f}" if projeto.valor else "R$ 0,00")
            ws.cell(row=row, column=4, value=projeto.data_inicio.strftime('%d/%m/%Y') if projeto.data_inicio else 'N/A')
            ws.cell(row=row, column=5, value=projeto.data_encerramento.strftime('%d/%m/%Y') if projeto.data_encerramento else 'N/A')
            ws.cell(row=row, column=6, value=situacao_map.get(projeto.situacao, 'N/A'))
        
        # Ajustar largura das colunas
        column_widths = {
            'A': 15,  # Código
            'B': 30,  # Nome 
            'C': 15,  # Valor
            'D': 15,  # Data Início
            'E': 15,  # Data Encerramento
            'F': 20   # Situação
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="relatorio_projetos.xlsx"'
        
        return response
    
    def generate_json_fallback(self):
        # Fallback JSON se Excel não funcionar
        projetos = Projeto.objects.all().order_by('-data_inicio')
        
        situacao_map = {
            '1': 'Aguardando Início', '2': 'Em Andamento', '3': 'Paralisado',
            '4': 'Suspenso', '5': 'Cancelado', '6': 'Concluído'
        }
        
        data = []
        for projeto in projetos:
            data.append({
                'codigo': projeto.cod_projeto,
                'nome': projeto.nome,
                'valor': float(projeto.valor) if projeto.valor else 0.0,
                'data_inicio': projeto.data_inicio.strftime('%d/%m/%Y') if projeto.data_inicio else None,
                'data_encerramento': projeto.data_encerramento.strftime('%d/%m/%Y') if projeto.data_encerramento else None,
                'situacao': situacao_map.get(projeto.situacao, 'N/A')
            })
        
        response = HttpResponse(
            json.dumps(data, indent=2, ensure_ascii=False),
            content_type='application/json'
        )
        response['Content-Disposition'] = 'attachment; filename="relatorio_projetos.json"'
        
        return response

class RelatorioContratosView(LoginRequiredMixin, TemplateView):
    template_name = 'relatorios/relatorio_contratos.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Dados completos de contratos
        contratos = Contrato.objects.all().order_by('-data_inicio')
        
        # Resumos estatísticos
        total_contratos = contratos.count()
        contratos_ativos = contratos.filter(situacao__in=['1', '2']).count()
        contratos_concluidos = contratos.filter(situacao='6').count()
        valor_total = contratos.aggregate(total=Sum('valor'))['total'] or 0
        
        # Contratos por tipo de pessoa
        contratos_pf = contratos.filter(tipo_pessoa=1).count()
        contratos_pj = contratos.filter(tipo_pessoa=2).count()
        
        context.update({
            'contratos': contratos,
            'total_contratos': total_contratos,
            'contratos_ativos': contratos_ativos,
            'contratos_concluidos': contratos_concluidos,
            'valor_total': valor_total,
            'contratos_pf': contratos_pf,
            'contratos_pj': contratos_pj,
        })
        
        return context

class RelatorioContratosGenerateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        format_type = request.GET.get('format', 'html')
        
        if format_type == 'pdf':
            return self.generate_pdf()
        elif format_type == 'excel':
            return self.generate_excel()
        else:
            return redirect('relatorios:relatorio_contratos')
    
    def generate_pdf(self):
        if not PDF_AVAILABLE:
            # Fallback para CSV se PDF não estiver disponível
            return self.generate_csv_fallback()
            
        if PDF_LIBRARY == 'fpdf':
            return self.generate_pdf_fpdf()
        else:
            # Fallback para CSV se não conseguir gerar PDF
            return self.generate_csv_fallback()
    
    def generate_pdf_fpdf(self):
        # Gerar PDF para relatório de contratos usando FPDF
        try:
            from fpdf import FPDF
            from fpdf.enums import XPos, YPos
            
            pdf = FPDF()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)
            
            # Título
            pdf.set_font('helvetica', 'B', 16)
            pdf.cell(0, 10, 'RELATÓRIO DE CONTRATOS - FUNETEC', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            
            # Data de geração
            pdf.set_font('helvetica', '', 10)
            pdf.cell(0, 5, f'Gerado em: {timezone.now().strftime("%d/%m/%Y às %H:%M")}', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            pdf.ln(10)
            
            # Cabeçalho da tabela
            pdf.set_font('helvetica', 'B', 8)
            pdf.set_fill_color(8, 149, 97)  # Verde FUNETEC
            pdf.set_text_color(255, 255, 255)  # Texto branco
            
            # Larguras das colunas
            col_widths = [20, 40, 30, 20, 25, 20, 20, 25]
            headers = ['Número', 'Contratado', 'CPF/CNPJ', 'Tipo', 'Valor', 'Data Início', 'Data Fim', 'Situação']
            
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 8, header, border=1, align='C', fill=True)
            pdf.ln()
            
            # Dados dos contratos
            contratos = Contrato.objects.all().order_by('-data_inicio')
            
            situacao_map = {
                '1': 'Aguardando Início', '2': 'Em Andamento', '3': 'Paralisado',
                '4': 'Suspenso', '5': 'Cancelado', '6': 'Concluído'
            }
            
            pdf.set_text_color(0, 0, 0)  # Texto preto
            pdf.set_font('helvetica', '', 7)
            
            for contrato in contratos:
                # Verificar se precisa de nova página
                if pdf.get_y() > 270:
                    pdf.add_page()
                    # Repetir cabeçalho
                    pdf.set_font('helvetica', 'B', 8)
                    pdf.set_fill_color(8, 149, 97)
                    pdf.set_text_color(255, 255, 255)
                    for i, header in enumerate(headers):
                        pdf.cell(col_widths[i], 8, header, border=1, align='C', fill=True)
                    pdf.ln()
                    pdf.set_text_color(0, 0, 0)
                    pdf.set_font('helvetica', '', 7)
                
                # Dados da linha
                contratado = (contrato.contratado or 'N/A')[:20] + '...' if len(contrato.contratado or '') > 20 else (contrato.contratado or 'N/A')
                cpf_cnpj = contrato.cpf_cnpj or 'N/A'
                tipo = 'PF' if contrato.tipo_pessoa == 1 else 'PJ' if contrato.tipo_pessoa == 2 else 'N/A'
                valor = f"R$ {contrato.valor:,.0f}" if contrato.valor else "R$ 0"
                data_inicio = contrato.data_inicio.strftime('%d/%m/%Y') if contrato.data_inicio else 'N/A'
                data_fim = contrato.data_fim.strftime('%d/%m/%Y') if contrato.data_fim else 'N/A'
                situacao = situacao_map.get(contrato.situacao, 'N/A')
                
                row_data = [contrato.num_contrato, contratado, cpf_cnpj, tipo, valor, data_inicio, data_fim, situacao]
                
                for i, data in enumerate(row_data):
                    align = 'R' if i == 4 else 'C'  # Valor alinhado à direita
                    pdf.cell(col_widths[i], 6, str(data), border=1, align=align)
                pdf.ln()
            
            # Gerar o PDF (método correto para fpdf2)
            pdf_output = pdf.output()
            if isinstance(pdf_output, (bytearray, bytes)):
                pdf_bytes = bytes(pdf_output)
            else:
                pdf_bytes = pdf_output.encode('latin-1')
            
            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="relatorio_contratos.pdf"'
            
            return response
            
        except Exception as e:
            # Se falhar a geração de PDF, usar fallback CSV
            return self.generate_csv_fallback()
    
    def generate_old_pdf(self):
        # Implementação antiga com reportlab (não funciona)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Elementos do documento
        elements = []
        
        # Título
        title = Paragraph("RELATÓRIO DE CONTRATOS - FUNETEC", title_style)
        elements.append(title)
        
        # Data de geração
        date_text = f"Gerado em: {timezone.now().strftime('%d/%m/%Y às %H:%M')}"
        date_para = Paragraph(date_text, styles['Normal'])
        elements.append(date_para)
        elements.append(Spacer(1, 20))
        
        # Dados dos contratos
        contratos = Contrato.objects.all().order_by('-data_inicio')
        
        situacao_map = {
            '1': 'Aguardando Início', '2': 'Em Andamento', '3': 'Paralisado',
            '4': 'Suspenso', '5': 'Cancelado', '6': 'Concluído'
        }
        
        # Tabela de dados
        data = [['Número', 'Contratado', 'CPF/CNPJ', 'Tipo', 'Valor', 'Data Início', 'Data Fim', 'Situação']]
        
        for contrato in contratos:
            data.append([
                contrato.num_contrato,
                contrato.contratado or 'N/A',
                contrato.cpf_cnpj or 'N/A',
                'Pessoa Física' if contrato.tipo_pessoa == 1 else 'Pessoa Jurídica' if contrato.tipo_pessoa == 2 else 'N/A',
                f"R$ {contrato.valor:,.2f}" if contrato.valor else "R$ 0,00",
                contrato.data_inicio.strftime('%d/%m/%Y') if contrato.data_inicio else 'N/A',
                contrato.data_fim.strftime('%d/%m/%Y') if contrato.data_fim else 'N/A',
                situacao_map.get(contrato.situacao, 'N/A')
            ])
        
        # Criar tabela
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#089561')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Construir PDF
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="relatorio_contratos.pdf"'
        
        return response
    
    def generate_csv_fallback(self):
        # Fallback CSV se PDF não funcionar
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="relatorio_contratos.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Número', 'Contratado', 'CPF/CNPJ', 'Tipo', 'Valor', 'Data Início', 'Data Fim', 'Situação'])
        
        contratos = Contrato.objects.all().order_by('-data_inicio')
        
        situacao_map = {
            '1': 'Aguardando Início', '2': 'Em Andamento', '3': 'Paralisado',
            '4': 'Suspenso', '5': 'Cancelado', '6': 'Concluído'
        }
        
        for contrato in contratos:
            writer.writerow([
                contrato.num_contrato,
                contrato.contratado or 'N/A',
                contrato.cpf_cnpj or 'N/A',
                'Pessoa Física' if contrato.tipo_pessoa == 1 else 'Pessoa Jurídica' if contrato.tipo_pessoa == 2 else 'N/A',
                f"R$ {contrato.valor:,.2f}" if contrato.valor else "R$ 0,00",
                contrato.data_inicio.strftime('%d/%m/%Y') if contrato.data_inicio else 'N/A',
                contrato.data_fim.strftime('%d/%m/%Y') if contrato.data_fim else 'N/A',
                situacao_map.get(contrato.situacao, 'N/A')
            ])
        
        return response
    
    def generate_excel(self):
        if not EXCEL_AVAILABLE:
            # Fallback para JSON se Excel não estiver disponível
            return self.generate_json_fallback()
            
        # Gerar Excel para relatório de contratos
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Contratos"
        
        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="089561", end_color="089561", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        # Título
        ws.merge_cells('A1:H1')
        ws['A1'] = "RELATÓRIO DE CONTRATOS - FUNETEC"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Data de geração
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Gerado em: {timezone.now().strftime('%d/%m/%Y às %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Cabeçalhos das colunas
        headers = ['Número', 'Contratado', 'CPF/CNPJ', 'Tipo', 'Valor', 'Data Início', 'Data Fim', 'Situação']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Dados dos contratos
        contratos = Contrato.objects.all().order_by('-data_inicio')
        
        situacao_map = {
            '1': 'Aguardando Início', '2': 'Em Andamento', '3': 'Paralisado',
            '4': 'Suspenso', '5': 'Cancelado', '6': 'Concluído'
        }
        
        for row, contrato in enumerate(contratos, 5):
            ws.cell(row=row, column=1, value=contrato.num_contrato)
            ws.cell(row=row, column=2, value=contrato.contratado or 'N/A')
            ws.cell(row=row, column=3, value=contrato.cpf_cnpj or 'N/A')
            ws.cell(row=row, column=4, value='Pessoa Física' if contrato.tipo_pessoa == 1 else 'Pessoa Jurídica' if contrato.tipo_pessoa == 2 else 'N/A')
            ws.cell(row=row, column=5, value=f"R$ {contrato.valor:,.2f}" if contrato.valor else "R$ 0,00")
            ws.cell(row=row, column=6, value=contrato.data_inicio.strftime('%d/%m/%Y') if contrato.data_inicio else 'N/A')
            ws.cell(row=row, column=7, value=contrato.data_fim.strftime('%d/%m/%Y') if contrato.data_fim else 'N/A')
            ws.cell(row=row, column=8, value=situacao_map.get(contrato.situacao, 'N/A'))
        
        # Ajustar largura das colunas
        column_widths = {
            'A': 15,  # Número
            'B': 25,  # Contratado
            'C': 20,  # CPF/CNPJ
            'D': 15,  # Tipo
            'E': 15,  # Valor
            'F': 15,  # Data Início
            'G': 15,  # Data Fim
            'H': 20   # Situação
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="relatorio_contratos.xlsx"'
        
        return response
    
    def generate_json_fallback(self):
        # Fallback JSON se Excel não funcionar
        contratos = Contrato.objects.all().order_by('-data_inicio')
        
        situacao_map = {
            '1': 'Aguardando Início', '2': 'Em Andamento', '3': 'Paralisado',
            '4': 'Suspenso', '5': 'Cancelado', '6': 'Concluído'
        }
        
        data = []
        for contrato in contratos:
            data.append({
                'numero': contrato.num_contrato,
                'contratado': contrato.contratado or 'N/A',
                'cpf_cnpj': contrato.cpf_cnpj or 'N/A',
                'tipo': 'Pessoa Física' if contrato.tipo_pessoa == 1 else 'Pessoa Jurídica' if contrato.tipo_pessoa == 2 else 'N/A',
                'valor': float(contrato.valor) if contrato.valor else 0.0,
                'data_inicio': contrato.data_inicio.strftime('%d/%m/%Y') if contrato.data_inicio else None,
                'data_fim': contrato.data_fim.strftime('%d/%m/%Y') if contrato.data_fim else None,
                'situacao': situacao_map.get(contrato.situacao, 'N/A')
            })
        
        response = HttpResponse(
            json.dumps(data, indent=2, ensure_ascii=False),
            content_type='application/json'
        )
        response['Content-Disposition'] = 'attachment; filename="relatorio_contratos.json"'
        
        return response

class RelatorioFinanceiroView(LoginRequiredMixin, TemplateView):
    template_name = 'relatorios/relatorio_financeiro.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Dados financeiros completos
        parcelas_pagas = ItemContrato.objects.filter(valor_pago__gt=0).select_related('num_contrato').order_by('-data_pagamento')
        todas_parcelas = ItemContrato.objects.all()
        
        # Resumos financeiros
        total_recebido = parcelas_pagas.aggregate(total=Sum('valor_pago'))['total'] or 0
        total_previsto = todas_parcelas.aggregate(total=Sum('valor_parcela'))['total'] or 0
        total_pendente = total_previsto - total_recebido
        
        # Parcelas por situação
        parcelas_liquidadas = todas_parcelas.filter(situacao='3').count()
        parcelas_vencidas = todas_parcelas.filter(
            data_vencimento__lt=timezone.now().date(),
            situacao__in=['1', '2']
        ).count()
        
        # Pagamentos por mês (ano atual) - SQLite compatible
        from django.db.models.functions import Extract
        pagamentos_mensais = parcelas_pagas.filter(
            data_pagamento__year=timezone.now().year
        ).annotate(
            month=Extract('data_pagamento', 'month')
        ).values('month').annotate(
            total=Sum('valor_pago'),
            count=Count('id')
        ).order_by('month')
        
        context.update({
            'parcelas_pagas': parcelas_pagas[:20],  # Últimas 20
            'total_recebido': total_recebido,
            'total_previsto': total_previsto,
            'total_pendente': total_pendente,
            'parcelas_liquidadas': parcelas_liquidadas,
            'parcelas_vencidas': parcelas_vencidas,
            'pagamentos_mensais': pagamentos_mensais,
            'taxa_recebimento': (total_recebido / total_previsto * 100) if total_previsto > 0 else 0,
        })
        
        return context

class RelatorioFinanceiroGenerateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        format_type = request.GET.get('format', 'html')
        
        if format_type == 'pdf':
            return self.generate_pdf()
        elif format_type == 'excel':
            return self.generate_excel()
        else:
            return redirect('relatorios:relatorio_financeiro')
    
    def generate_pdf(self):
        if not PDF_AVAILABLE:
            # Fallback para CSV se PDF não estiver disponível
            return self.generate_csv_fallback()
            
        if PDF_LIBRARY == 'fpdf':
            return self.generate_pdf_fpdf()
        else:
            # Fallback para CSV se não conseguir gerar PDF
            return self.generate_csv_fallback()
    
    def generate_pdf_fpdf(self):
        # Gerar PDF para relatório financeiro usando FPDF
        try:
            from fpdf import FPDF
            from fpdf.enums import XPos, YPos
            
            pdf = FPDF()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)
            
            # Título
            pdf.set_font('helvetica', 'B', 16)
            pdf.cell(0, 10, 'RELATÓRIO FINANCEIRO - FUNETEC', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            
            # Data de geração
            pdf.set_font('helvetica', '', 10)
            pdf.cell(0, 5, f'Gerado em: {timezone.now().strftime("%d/%m/%Y às %H:%M")}', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            pdf.ln(15)
            
            # Resumo financeiro
            parcelas_pagas = ItemContrato.objects.filter(valor_pago__gt=0).select_related('num_contrato')
            todas_parcelas = ItemContrato.objects.all()
            
            total_recebido = parcelas_pagas.aggregate(total=Sum('valor_pago'))['total'] or 0
            total_previsto = todas_parcelas.aggregate(total=Sum('valor_parcela'))['total'] or 0
            total_pendente = total_previsto - total_recebido
            taxa_recebimento = (total_recebido / total_previsto * 100) if total_previsto > 0 else 0
            
            # Caixa de resumo
            pdf.set_fill_color(248, 249, 250)  # Cinza claro
            pdf.rect(10, pdf.get_y(), 190, 40, 'F')
            
            pdf.set_font('helvetica', 'B', 12)
            pdf.cell(0, 8, 'RESUMO FINANCEIRO', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            
            pdf.set_font('helvetica', '', 10)
            pdf.cell(95, 6, f'Total Recebido: R$ {total_recebido:,.2f}', border=0, align='L')
            pdf.cell(95, 6, f'Total Previsto: R$ {total_previsto:,.2f}', border=0, align='L')
            pdf.ln()
            pdf.cell(95, 6, f'Total Pendente: R$ {total_pendente:,.2f}', border=0, align='L')
            pdf.cell(95, 6, f'Taxa de Recebimento: {taxa_recebimento:.1f}%', border=0, align='L')
            pdf.ln(15)
            
            # Cabeçalho da tabela
            pdf.set_font('helvetica', 'B', 9)
            pdf.set_fill_color(8, 149, 97)  # Verde FUNETEC
            pdf.set_text_color(255, 255, 255)  # Texto branco
            
            # Larguras das colunas
            col_widths = [25, 25, 25, 25, 25, 25, 40]
            headers = ['Data', 'Contrato', 'Parcela', 'Valor Pago', 'Valor Total', 'Diferença', 'Status']
            
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 8, header, border=1, align='C', fill=True)
            pdf.ln()
            
            # Dados das parcelas pagas
            parcelas = parcelas_pagas.order_by('-data_pagamento')[:50]  # Últimas 50
            
            situacao_map = {
                '1': 'Pendente', '2': 'Processando', '3': 'Liquidada', '4': 'Cancelada'
            }
            
            pdf.set_text_color(0, 0, 0)  # Texto preto
            pdf.set_font('helvetica', '', 8)
            
            for parcela in parcelas:
                # Verificar se precisa de nova página
                if pdf.get_y() > 270:
                    pdf.add_page()
                    # Repetir cabeçalho
                    pdf.set_font('helvetica', 'B', 9)
                    pdf.set_fill_color(8, 149, 97)
                    pdf.set_text_color(255, 255, 255)
                    for i, header in enumerate(headers):
                        pdf.cell(col_widths[i], 8, header, border=1, align='C', fill=True)
                    pdf.ln()
                    pdf.set_text_color(0, 0, 0)
                    pdf.set_font('helvetica', '', 8)
                
                # Dados da linha
                data_pag = parcela.data_pagamento.strftime('%d/%m/%Y') if parcela.data_pagamento else 'N/A'
                contrato = parcela.num_contrato.num_contrato if parcela.num_contrato else 'N/A'
                num_parcela = f"{parcela.num_parcela}ª"
                valor_pago = f"R$ {parcela.valor_pago:,.0f}"
                valor_total = f"R$ {parcela.valor_parcela:,.0f}"
                diferenca = parcela.valor_parcela - parcela.valor_pago if parcela.valor_parcela and parcela.valor_pago else 0
                diferenca_str = f"R$ {diferenca:,.0f}"
                status = situacao_map.get(parcela.situacao, 'N/A')
                
                row_data = [data_pag, contrato, num_parcela, valor_pago, valor_total, diferenca_str, status]
                
                for i, data in enumerate(row_data):
                    align = 'R' if i in [3, 4, 5] else 'C'  # Valores alinhados à direita
                    pdf.cell(col_widths[i], 6, str(data), border=1, align=align)
                pdf.ln()
            
            # Gerar o PDF (método correto para fpdf2)
            pdf_output = pdf.output()
            if isinstance(pdf_output, (bytearray, bytes)):
                pdf_bytes = bytes(pdf_output)
            else:
                pdf_bytes = pdf_output.encode('latin-1')
            
            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="relatorio_financeiro.pdf"'
            
            return response
            
        except Exception as e:
            # Se falhar a geração de PDF, usar fallback CSV
            return self.generate_csv_fallback()
    
    def generate_old_pdf(self):
        # Implementação antiga com reportlab (não funciona)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Elementos do documento
        elements = []
        
        # Título
        title = Paragraph("RELATÓRIO FINANCEIRO - FUNETEC", title_style)
        elements.append(title)
        
        # Data de geração
        date_text = f"Gerado em: {timezone.now().strftime('%d/%m/%Y às %H:%M')}"
        date_para = Paragraph(date_text, styles['Normal'])
        elements.append(date_para)
        elements.append(Spacer(1, 20))
        
        # Resumo financeiro
        parcelas_pagas = ItemContrato.objects.filter(valor_pago__gt=0)
        todas_parcelas = ItemContrato.objects.all()
        
        total_recebido = parcelas_pagas.aggregate(total=Sum('valor_pago'))['total'] or 0
        total_previsto = todas_parcelas.aggregate(total=Sum('valor_parcela'))['total'] or 0
        total_pendente = total_previsto - total_recebido
        
        resumo_data = [
            ['RESUMO FINANCEIRO', '', '', ''],
            ['Total Recebido', f'R$ {total_recebido:,.2f}', '', ''],
            ['Total Previsto', f'R$ {total_previsto:,.2f}', '', ''],
            ['Total Pendente', f'R$ {total_pendente:,.2f}', '', ''],
            ['Taxa de Recebimento', f'{(total_recebido/total_previsto*100) if total_previsto > 0 else 0:.1f}%', '', '']
        ]
        
        resumo_table = Table(resumo_data)
        resumo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#089561')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(resumo_table)
        elements.append(Spacer(1, 20))
        
        # Tabela de pagamentos
        data = [['Data', 'Contrato', 'Parcela', 'Valor Pago', 'Valor Total', 'Status']]
        
        parcelas_pagas_list = parcelas_pagas.order_by('-data_pagamento')
        
        for parcela in parcelas_pagas_list:
            data.append([
                parcela.data_pagamento.strftime('%d/%m/%Y') if parcela.data_pagamento else 'N/A',
                str(parcela.num_contrato.num_contrato) if parcela.num_contrato else 'N/A',
                f"{parcela.num_parcela}ª parcela",
                f"R$ {parcela.valor_pago:,.2f}",
                f"R$ {parcela.valor_parcela:,.2f}",
                'Liquidado' if parcela.situacao == '3' else 'Pago Parcial'
            ])
        
        # Criar tabela
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#089561')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Construir PDF
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="relatorio_financeiro.pdf"'
        
        return response
    
    def generate_csv_fallback(self):
        # Fallback CSV se PDF não funcionar
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="relatorio_financeiro.csv"'
        
        writer = csv.writer(response)
        
        # Cabeçalho do resumo
        parcelas_pagas = ItemContrato.objects.filter(valor_pago__gt=0)
        todas_parcelas = ItemContrato.objects.all()
        
        total_recebido = parcelas_pagas.aggregate(total=Sum('valor_pago'))['total'] or 0
        total_previsto = todas_parcelas.aggregate(total=Sum('valor_parcela'))['total'] or 0
        total_pendente = total_previsto - total_recebido
        
        writer.writerow(['RELATÓRIO FINANCEIRO - FUNETEC'])
        writer.writerow([f'Gerado em: {timezone.now().strftime("%d/%m/%Y às %H:%M")}'])
        writer.writerow([])
        writer.writerow(['RESUMO'])
        writer.writerow(['Total Recebido', f'R$ {total_recebido:,.2f}'])
        writer.writerow(['Total Previsto', f'R$ {total_previsto:,.2f}'])
        writer.writerow(['Total Pendente', f'R$ {total_pendente:,.2f}'])
        writer.writerow(['Taxa de Recebimento', f'{(total_recebido/total_previsto*100) if total_previsto > 0 else 0:.1f}%'])
        writer.writerow([])
        
        # Cabeçalho da tabela
        writer.writerow(['Data', 'Contrato', 'Parcela', 'Valor Pago', 'Valor Total', 'Status'])
        
        # Dados dos pagamentos
        parcelas_pagas_list = parcelas_pagas.order_by('-data_pagamento')
        
        for parcela in parcelas_pagas_list:
            writer.writerow([
                parcela.data_pagamento.strftime('%d/%m/%Y') if parcela.data_pagamento else 'N/A',
                str(parcela.num_contrato.num_contrato) if parcela.num_contrato else 'N/A',
                f"{parcela.num_parcela}ª parcela",
                f"R$ {parcela.valor_pago:,.2f}",
                f"R$ {parcela.valor_parcela:,.2f}",
                'Liquidado' if parcela.situacao == '3' else 'Pago Parcial'
            ])
        
        return response
    
    def generate_excel(self):
        if not EXCEL_AVAILABLE:
            # Fallback para JSON se Excel não estiver disponível
            return self.generate_json_fallback()
            
        # Gerar Excel para relatório financeiro
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório Financeiro"
        
        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="089561", end_color="089561", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        # Título
        ws.merge_cells('A1:F1')
        ws['A1'] = "RELATÓRIO FINANCEIRO - FUNETEC"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Data de geração
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Gerado em: {timezone.now().strftime('%d/%m/%Y às %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Resumo financeiro
        parcelas_pagas = ItemContrato.objects.filter(valor_pago__gt=0)
        todas_parcelas = ItemContrato.objects.all()
        
        total_recebido = parcelas_pagas.aggregate(total=Sum('valor_pago'))['total'] or 0
        total_previsto = todas_parcelas.aggregate(total=Sum('valor_parcela'))['total'] or 0
        total_pendente = total_previsto - total_recebido
        
        # Resumo de dados
        ws['A4'] = "RESUMO FINANCEIRO"
        ws['A4'].font = Font(bold=True, size=12)
        ws['A5'] = f"Total Recebido: R$ {total_recebido:,.2f}"
        ws['A6'] = f"Total Previsto: R$ {total_previsto:,.2f}" 
        ws['A7'] = f"Total Pendente: R$ {total_pendente:,.2f}"
        ws['A8'] = f"Taxa de Recebimento: {(total_recebido/total_previsto*100) if total_previsto > 0 else 0:.1f}%"
        
        # Cabeçalhos das colunas (linha 10)
        headers = ['Data', 'Contrato', 'Parcela', 'Valor Pago', 'Valor Total', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Dados dos pagamentos
        parcelas_pagas_list = parcelas_pagas.order_by('-data_pagamento')
        
        for row, parcela in enumerate(parcelas_pagas_list, 11):
            ws.cell(row=row, column=1, value=parcela.data_pagamento.strftime('%d/%m/%Y') if parcela.data_pagamento else 'N/A')
            ws.cell(row=row, column=2, value=str(parcela.num_contrato.num_contrato) if parcela.num_contrato else 'N/A')
            ws.cell(row=row, column=3, value=f"{parcela.num_parcela}ª parcela")
            ws.cell(row=row, column=4, value=f"R$ {parcela.valor_pago:,.2f}")
            ws.cell(row=row, column=5, value=f"R$ {parcela.valor_parcela:,.2f}")
            ws.cell(row=row, column=6, value='Liquidado' if parcela.situacao == '3' else 'Pago Parcial')
        
        # Ajustar largura das colunas
        column_widths = {
            'A': 15,  # Data
            'B': 15,  # Contrato
            'C': 15,  # Parcela
            'D': 15,  # Valor Pago
            'E': 15,  # Valor Total
            'F': 15   # Status
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="relatorio_financeiro.xlsx"'
        
        return response
    
    def generate_json_fallback(self):
        # Fallback JSON se Excel não funcionar - Gerar JSON para relatório financeiro
        parcelas_pagas = ItemContrato.objects.filter(valor_pago__gt=0)
        todas_parcelas = ItemContrato.objects.all()
        
        total_recebido = parcelas_pagas.aggregate(total=Sum('valor_pago'))['total'] or 0
        total_previsto = todas_parcelas.aggregate(total=Sum('valor_parcela'))['total'] or 0
        total_pendente = total_previsto - total_recebido
        
        # Resumo
        resumo = {
            'total_recebido': float(total_recebido) if total_recebido else 0.0,
            'total_previsto': float(total_previsto) if total_previsto else 0.0,
            'total_pendente': float(total_pendente) if total_pendente else 0.0,
            'taxa_recebimento': float((total_recebido/total_previsto*100) if total_previsto > 0 else 0)
        }
        
        # Pagamentos
        pagamentos = []
        parcelas_pagas_list = parcelas_pagas.order_by('-data_pagamento')
        
        for parcela in parcelas_pagas_list:
            pagamentos.append({
                'data': parcela.data_pagamento.strftime('%d/%m/%Y') if parcela.data_pagamento else None,
                'contrato': str(parcela.num_contrato.num_contrato) if parcela.num_contrato else 'N/A',
                'parcela': f"{parcela.num_parcela}ª parcela",
                'valor_pago': float(parcela.valor_pago) if parcela.valor_pago else 0.0,
                'valor_total': float(parcela.valor_parcela) if parcela.valor_parcela else 0.0,
                'status': 'Liquidado' if parcela.situacao == '3' else 'Pago Parcial'
            })
        
        data = {
            'relatorio': 'Financeiro - FUNETEC',
            'gerado_em': timezone.now().strftime('%d/%m/%Y às %H:%M'),
            'resumo': resumo,
            'pagamentos': pagamentos
        }
        
        response = HttpResponse(
            json.dumps(data, indent=2, ensure_ascii=False),
            content_type='application/json'
        )
        response['Content-Disposition'] = 'attachment; filename="relatorio_financeiro.json"'
        
        return response
    
class RelatorioCustomView(LoginRequiredMixin, TemplateView):
    template_name = 'relatorios/relatorio_custom.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Importar formulário
        from apps.relatorios.forms import RelatorioCustomForm
        
        # Formulário de filtros
        form = RelatorioCustomForm(self.request.GET or None)
        context['form'] = form
        
        # Se o formulário foi submetido e é válido, processar dados
        if form.is_valid():
            context['results'] = self.get_filtered_results(form.cleaned_data)
            context['show_results'] = True
        else:
            context['show_results'] = False
        
        return context
    
    def get_filtered_results(self, filters):
        """Obter resultados filtrados baseado no formulário"""
        tipo_relatorio = filters.get('tipo_relatorio')
        
        if tipo_relatorio == 'projetos':
            return self.get_projetos_filtered(filters)
        elif tipo_relatorio == 'contratos':
            return self.get_contratos_filtered(filters)
        elif tipo_relatorio == 'financeiro':
            return self.get_financeiro_filtered(filters)
        elif tipo_relatorio == 'mixto':
            return self.get_mixto_filtered(filters)
        
        return {}
    
    def get_projetos_filtered(self, filters):
        """Filtrar projetos baseado nos critérios"""
        queryset = Projeto.objects.all()
        
        # Filtros de data
        if filters.get('data_inicio'):
            queryset = queryset.filter(data_inicio__gte=filters['data_inicio'])
        if filters.get('data_fim'):
            queryset = queryset.filter(data_inicio__lte=filters['data_fim'])
        
        # Filtros de valor
        if filters.get('valor_minimo'):
            queryset = queryset.filter(valor__gte=filters['valor_minimo'])
        if filters.get('valor_maximo'):
            queryset = queryset.filter(valor__lte=filters['valor_maximo'])
        
        # Filtro de situação
        if filters.get('situacao'):
            queryset = queryset.filter(situacao=filters['situacao'])
        
        # Filtro de nome
        if filters.get('nome_projeto'):
            queryset = queryset.filter(nome__icontains=filters['nome_projeto'])
        
        # Ordenação
        ordem = filters.get('ordenacao', 'nome') or 'nome'
        if ordem and filters.get('ordem_desc'):
            ordem = f'-{ordem}'
        if ordem:
            queryset = queryset.order_by(ordem)
        else:
            queryset = queryset.order_by('nome')
        
        # Totalizadores
        total_projetos = queryset.count()
        valor_total = queryset.aggregate(total=Sum('valor'))['total'] or 0
        
        return {
            'tipo': 'projetos',
            'dados': queryset[:100],  # Limite de 100 registros
            'total_registros': total_projetos,
            'valor_total': valor_total,
            'incluir_totais': filters.get('incluir_totais', True)
        }
    
    def get_contratos_filtered(self, filters):
        """Filtrar contratos baseado nos critérios"""
        queryset = Contrato.objects.all()
        
        # Filtros de data
        if filters.get('data_inicio'):
            queryset = queryset.filter(data_inicio__gte=filters['data_inicio'])
        if filters.get('data_fim'):
            queryset = queryset.filter(data_inicio__lte=filters['data_fim'])
        
        # Filtros de valor
        if filters.get('valor_minimo'):
            queryset = queryset.filter(valor__gte=filters['valor_minimo'])
        if filters.get('valor_maximo'):
            queryset = queryset.filter(valor__lte=filters['valor_maximo'])
        
        # Filtro de situação
        if filters.get('situacao'):
            queryset = queryset.filter(situacao=filters['situacao'])
        
        # Filtro de tipo de pessoa
        if filters.get('tipo_pessoa'):
            queryset = queryset.filter(tipo_pessoa=filters['tipo_pessoa'])
        
        # Filtro de contratado
        if filters.get('contratado'):
            queryset = queryset.filter(contratado__icontains=filters['contratado'])
        
        # Ordenação - mapear campos para contratos
        ordem_original = filters.get('ordenacao', 'nome') or 'nome'
        ordem_map = {
            'nome': 'contratado',
            'data_fim': 'data_fim',
            'data_inicio': 'data_inicio'
        }
        
        # Mapear a ordenação
        ordem = ordem_map.get(ordem_original, ordem_original)
        
        # Aplicar ordenação descendente se solicitado
        if filters.get('ordem_desc'):
            ordem = f'-{ordem}'
        
        # Garantir que sempre há uma ordenação válida
        if not ordem or ordem == '-':
            ordem = 'contratado'
        
        queryset = queryset.order_by(ordem)
        
        # Totalizadores
        total_contratos = queryset.count()
        valor_total = queryset.aggregate(total=Sum('valor'))['total'] or 0
        
        return {
            'tipo': 'contratos',
            'dados': queryset[:100],
            'total_registros': total_contratos,
            'valor_total': valor_total,
            'incluir_totais': filters.get('incluir_totais', True)
        }
    
    def get_financeiro_filtered(self, filters):
        """Filtrar dados financeiros baseado nos critérios"""
        queryset = ItemContrato.objects.filter(valor_pago__gt=0).select_related('num_contrato')
        
        # Filtros de data (usando data_pagamento)
        if filters.get('data_inicio'):
            queryset = queryset.filter(data_pagamento__gte=filters['data_inicio'])
        if filters.get('data_fim'):
            queryset = queryset.filter(data_pagamento__lte=filters['data_fim'])
        
        # Filtros de valor (usando valor_pago)
        if filters.get('valor_minimo'):
            queryset = queryset.filter(valor_pago__gte=filters['valor_minimo'])
        if filters.get('valor_maximo'):
            queryset = queryset.filter(valor_pago__lte=filters['valor_maximo'])
        
        # Filtro por contratado
        if filters.get('contratado'):
            queryset = queryset.filter(num_contrato__contratado__icontains=filters['contratado'])
        
        # Ordenação
        ordem = filters.get('ordenacao', 'data_pagamento') or 'data_pagamento'
        ordem_map = {
            'nome': 'num_contrato__contratado',
            'valor': 'valor_pago',
            'data_inicio': 'data_pagamento',
            'data_fim': 'data_vencimento'
        }
        ordem = ordem_map.get(ordem, 'data_pagamento')
        if filters.get('ordem_desc'):
            ordem = f'-{ordem}'
        
        if not ordem:
            ordem = 'data_pagamento'
        
        queryset = queryset.order_by(ordem)
        
        # Totalizadores
        total_parcelas = queryset.count()
        valor_total_pago = queryset.aggregate(total=Sum('valor_pago'))['total'] or 0
        valor_total_previsto = queryset.aggregate(total=Sum('valor_parcela'))['total'] or 0
        
        return {
            'tipo': 'financeiro',
            'dados': queryset[:100],
            'total_registros': total_parcelas,
            'valor_total_pago': valor_total_pago,
            'valor_total_previsto': valor_total_previsto,
            'incluir_totais': filters.get('incluir_totais', True)
        }
    
    def get_mixto_filtered(self, filters):
        """Combinar dados de projetos e contratos"""
        projetos = self.get_projetos_filtered(filters)
        contratos = self.get_contratos_filtered(filters)
        
        return {
            'tipo': 'mixto',
            'projetos': projetos,
            'contratos': contratos,
            'incluir_totais': filters.get('incluir_totais', True)
        }


class RelatorioCustomGenerateView(LoginRequiredMixin, View):
    """View para gerar PDF ou Excel de relatórios customizados"""
    
    def get(self, request, *args, **kwargs):
        # Importar formulário
        from apps.relatorios.forms import RelatorioCustomForm
        
        # Validar formulário
        form = RelatorioCustomForm(request.GET)
        if not form.is_valid():
            messages.error(request, 'Filtros inválidos. Verifique os dados inseridos.')
            return redirect('relatorios:relatorio_custom')
        
        # Obter dados filtrados usando a mesma lógica da view principal
        custom_view = RelatorioCustomView()
        custom_view.request = request
        results = custom_view.get_filtered_results(form.cleaned_data)
        
        # Verificar formato
        format_type = request.GET.get('format', 'pdf')
        
        if format_type == 'pdf':
            return self.generate_pdf(results, form.cleaned_data)
        elif format_type == 'excel':
            return self.generate_excel(results, form.cleaned_data)
        else:
            messages.error(request, 'Formato inválido.')
            return redirect('relatorios:relatorio_custom')
    
    def generate_pdf(self, results, filters):
        """Gerar PDF customizado"""
        from fpdf import FPDF
        import os
        
        pdf = FPDF()
        pdf.add_page()
        
        # Configurar fontes
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, 'FUNETEC - Relatório Customizado', 0, 1, 'C')
        pdf.ln(5)
        
        # Informações do filtro
        pdf.set_font('Arial', '', 10)
        from apps.relatorios.forms import RelatorioCustomForm
        pdf.cell(0, 8, f'Tipo: {dict(RelatorioCustomForm.TIPO_CHOICES).get(filters.get("tipo_relatorio"), "N/A")}', 0, 1)
        pdf.cell(0, 8, f'Gerado em: {timezone.now().strftime("%d/%m/%Y às %H:%M")}', 0, 1)
        
        if filters.get('data_inicio') and filters.get('data_fim'):
            pdf.cell(0, 8, f'Período: {filters["data_inicio"].strftime("%d/%m/%Y")} a {filters["data_fim"].strftime("%d/%m/%Y")}', 0, 1)
        
        pdf.ln(5)
        
        # Conteúdo baseado no tipo
        if results.get('tipo') == 'projetos':
            self._add_projetos_to_pdf(pdf, results)
        elif results.get('tipo') == 'contratos':
            self._add_contratos_to_pdf(pdf, results)
        elif results.get('tipo') == 'financeiro':
            self._add_financeiro_to_pdf(pdf, results)
        elif results.get('tipo') == 'mixto':
            self._add_mixto_to_pdf(pdf, results)
        
        # Totalizadores
        if results.get('incluir_totais'):
            self._add_totals_to_pdf(pdf, results)
        
        # Gerar o PDF
        pdf_output = pdf.output()
        if isinstance(pdf_output, (bytearray, bytes)):
            pdf_bytes = bytes(pdf_output)
        else:
            pdf_bytes = pdf_output.encode('latin-1')
        
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="relatorio_customizado.pdf"'
        return response
    
    def _add_projetos_to_pdf(self, pdf, results):
        """Adicionar projetos ao PDF"""
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, 'PROJETOS', 0, 1)
        pdf.ln(2)
        
        # Cabeçalho da tabela
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(25, 8, 'Código', 1, 0, 'C')
        pdf.cell(60, 8, 'Nome', 1, 0, 'C')
        pdf.cell(30, 8, 'Valor', 1, 0, 'C')
        pdf.cell(25, 8, 'Início', 1, 0, 'C')
        pdf.cell(30, 8, 'Situação', 1, 1, 'C')
        
        # Dados
        pdf.set_font('Arial', '', 8)
        for projeto in results.get('dados', []):
            pdf.cell(25, 6, str(projeto.cod_projeto), 1, 0)
            pdf.cell(60, 6, projeto.nome[:25] + '...' if len(projeto.nome) > 25 else projeto.nome, 1, 0)
            pdf.cell(30, 6, f'R$ {projeto.valor:.2f}' if projeto.valor is not None else 'R$ 0,00', 1, 0)
            pdf.cell(25, 6, projeto.data_inicio.strftime('%d/%m/%Y') if projeto.data_inicio else 'N/A', 1, 0)
            pdf.cell(30, 6, projeto.get_situacao_display() or 'N/A', 1, 1)
    
    def _add_contratos_to_pdf(self, pdf, results):
        """Adicionar contratos ao PDF"""
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, 'CONTRATOS', 0, 1)
        pdf.ln(2)
        
        # Cabeçalho da tabela
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(25, 8, 'Número', 1, 0, 'C')
        pdf.cell(50, 8, 'Contratado', 1, 0, 'C')
        pdf.cell(20, 8, 'Tipo', 1, 0, 'C')
        pdf.cell(30, 8, 'Valor', 1, 0, 'C')
        pdf.cell(25, 8, 'Início', 1, 0, 'C')
        pdf.cell(20, 8, 'Situação', 1, 1, 'C')
        
        # Dados
        pdf.set_font('Arial', '', 8)
        for contrato in results.get('dados', []):
            pdf.cell(25, 6, str(contrato.num_contrato), 1, 0)
            pdf.cell(50, 6, (contrato.contratado or 'N/A')[:20] + '...' if len(contrato.contratado or '') > 20 else (contrato.contratado or 'N/A'), 1, 0)
            tipo = 'PF' if contrato.tipo_pessoa == 1 else 'PJ' if contrato.tipo_pessoa == 2 else 'N/A'
            pdf.cell(20, 6, tipo, 1, 0)
            pdf.cell(30, 6, f'R$ {contrato.valor:.2f}' if contrato.valor is not None else 'R$ 0,00', 1, 0)
            pdf.cell(25, 6, contrato.data_inicio.strftime('%d/%m/%Y') if contrato.data_inicio else 'N/A', 1, 0)
            pdf.cell(20, 6, (contrato.get_situacao_display() or 'N/A')[:10], 1, 1)
    
    def _add_financeiro_to_pdf(self, pdf, results):
        """Adicionar dados financeiros ao PDF"""
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, 'RELATÓRIO FINANCEIRO', 0, 1)
        pdf.ln(2)
        
        # Cabeçalho da tabela
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(25, 8, 'Data Pag.', 1, 0, 'C')
        pdf.cell(25, 8, 'Contrato', 1, 0, 'C')
        pdf.cell(40, 8, 'Contratado', 1, 0, 'C')
        pdf.cell(15, 8, 'Parc.', 1, 0, 'C')
        pdf.cell(30, 8, 'Valor Pago', 1, 0, 'C')
        pdf.cell(25, 8, 'Status', 1, 1, 'C')
        
        # Dados
        pdf.set_font('Arial', '', 8)
        for parcela in results.get('dados', []):
            pdf.cell(25, 6, parcela.data_pagamento.strftime('%d/%m/%Y') if parcela.data_pagamento else 'N/A', 1, 0)
            pdf.cell(25, 6, str(parcela.num_contrato.num_contrato), 1, 0)
            contratado = (parcela.num_contrato.contratado or 'N/A')[:15]
            pdf.cell(40, 6, contratado, 1, 0)
            pdf.cell(15, 6, f'{parcela.num_parcela}ª', 1, 0)
            pdf.cell(30, 6, f'R$ {parcela.valor_pago:.2f}' if parcela.valor_pago is not None else 'R$ 0,00', 1, 0)
            pdf.cell(25, 6, (parcela.get_situacao_display() or 'N/A')[:10], 1, 1)
    
    def _add_mixto_to_pdf(self, pdf, results):
        """Adicionar relatório misto ao PDF"""
        # Projetos
        if results.get('projetos', {}).get('dados'):
            self._add_projetos_to_pdf(pdf, results['projetos'])
            pdf.ln(5)
        
        # Contratos
        if results.get('contratos', {}).get('dados'):
            self._add_contratos_to_pdf(pdf, results['contratos'])
    
    def _add_totals_to_pdf(self, pdf, results):
        """Adicionar totalizadores ao PDF"""
        pdf.ln(5)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 8, 'RESUMO', 0, 1)
        pdf.set_font('Arial', '', 9)
        
        if results.get('tipo') == 'projetos':
            pdf.cell(0, 6, f'Total de Projetos: {results.get("total_registros", 0)}', 0, 1)
            pdf.cell(0, 6, f'Valor Total: R$ {results.get("valor_total", 0):.2f}', 0, 1)
        elif results.get('tipo') == 'contratos':
            pdf.cell(0, 6, f'Total de Contratos: {results.get("total_registros", 0)}', 0, 1)
            pdf.cell(0, 6, f'Valor Total: R$ {results.get("valor_total", 0):.2f}', 0, 1)
        elif results.get('tipo') == 'financeiro':
            pdf.cell(0, 6, f'Total de Parcelas: {results.get("total_registros", 0)}', 0, 1)
            pdf.cell(0, 6, f'Valor Pago: R$ {results.get("valor_total_pago", 0):.2f}', 0, 1)
            pdf.cell(0, 6, f'Valor Previsto: R$ {results.get("valor_total_previsto", 0):.2f}', 0, 1)
        elif results.get('tipo') == 'mixto':
            projetos = results.get('projetos', {})
            contratos = results.get('contratos', {})
            pdf.cell(0, 6, f'Projetos: {projetos.get("total_registros", 0)} (R$ {projetos.get("valor_total", 0):.2f})', 0, 1)
            pdf.cell(0, 6, f'Contratos: {contratos.get("total_registros", 0)} (R$ {contratos.get("valor_total", 0):.2f})', 0, 1)
    
    def generate_excel(self, results, filters):
        """Gerar Excel customizado"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório Customizado"
        
        # Cabeçalho
        ws['A1'] = 'FUNETEC - Relatório Customizado'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        
        # Informações do filtro
        row = 3
        from apps.relatorios.forms import RelatorioCustomForm
        ws[f'A{row}'] = f'Tipo: {dict(RelatorioCustomForm.TIPO_CHOICES).get(filters.get("tipo_relatorio"), "N/A")}'
        row += 1
        ws[f'A{row}'] = f'Gerado em: {timezone.now().strftime("%d/%m/%Y às %H:%M")}'
        row += 2
        
        # Conteúdo baseado no tipo
        if results.get('tipo') == 'projetos':
            row = self._add_projetos_to_excel(ws, results, row)
        elif results.get('tipo') == 'contratos':
            row = self._add_contratos_to_excel(ws, results, row)
        elif results.get('tipo') == 'financeiro':
            row = self._add_financeiro_to_excel(ws, results, row)
        elif results.get('tipo') == 'mixto':
            row = self._add_mixto_to_excel(ws, results, row)
        
        # Totalizadores
        if results.get('incluir_totais'):
            self._add_totals_to_excel(ws, results, row)
        
        # Salvar em bytes
        from io import BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = HttpResponse(
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="relatorio_customizado.xlsx"'
        return response
    
    def _add_projetos_to_excel(self, ws, results, start_row):
        """Adicionar projetos ao Excel"""
        row = start_row
        
        # Título
        ws[f'A{row}'] = 'PROJETOS'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Cabeçalhos
        headers = ['Código', 'Nome', 'Valor', 'Data Início', 'Data Encerramento', 'Situação']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        row += 1
        
        # Dados
        for projeto in results.get('dados', []):
            ws.cell(row=row, column=1, value=projeto.cod_projeto)
            ws.cell(row=row, column=2, value=projeto.nome)
            ws.cell(row=row, column=3, value=float(projeto.valor) if projeto.valor is not None else 0.0)
            ws.cell(row=row, column=4, value=projeto.data_inicio.strftime('%d/%m/%Y') if projeto.data_inicio else 'N/A')
            ws.cell(row=row, column=5, value=projeto.data_encerramento.strftime('%d/%m/%Y') if projeto.data_encerramento else 'N/A')
            ws.cell(row=row, column=6, value=projeto.get_situacao_display())
            row += 1
        
        return row + 2
    
    def _add_contratos_to_excel(self, ws, results, start_row):
        """Adicionar contratos ao Excel"""
        row = start_row
        
        # Título
        ws[f'A{row}'] = 'CONTRATOS'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Cabeçalhos
        headers = ['Número', 'Contratado', 'CPF/CNPJ', 'Tipo', 'Valor', 'Data Início', 'Situação']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        row += 1
        
        # Dados
        for contrato in results.get('dados', []):
            ws.cell(row=row, column=1, value=contrato.num_contrato)
            ws.cell(row=row, column=2, value=contrato.contratado or 'N/A')
            ws.cell(row=row, column=3, value=contrato.cpf_cnpj or 'N/A')
            tipo = 'Pessoa Física' if contrato.tipo_pessoa == 1 else 'Pessoa Jurídica' if contrato.tipo_pessoa == 2 else 'N/A'
            ws.cell(row=row, column=4, value=tipo)
            ws.cell(row=row, column=5, value=float(contrato.valor) if contrato.valor is not None else 0.0)
            ws.cell(row=row, column=6, value=contrato.data_inicio.strftime('%d/%m/%Y') if contrato.data_inicio else 'N/A')
            ws.cell(row=row, column=7, value=contrato.get_situacao_display())
            row += 1
        
        return row + 2
    
    def _add_financeiro_to_excel(self, ws, results, start_row):
        """Adicionar dados financeiros ao Excel"""
        row = start_row
        
        # Título
        ws[f'A{row}'] = 'RELATÓRIO FINANCEIRO'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Cabeçalhos
        headers = ['Data Pagamento', 'Contrato', 'Contratado', 'Parcela', 'Valor Pago', 'Valor Total', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        row += 1
        
        # Dados
        for parcela in results.get('dados', []):
            ws.cell(row=row, column=1, value=parcela.data_pagamento.strftime('%d/%m/%Y') if parcela.data_pagamento else 'N/A')
            ws.cell(row=row, column=2, value=parcela.num_contrato.num_contrato)
            ws.cell(row=row, column=3, value=parcela.num_contrato.contratado or 'N/A')
            ws.cell(row=row, column=4, value=f'{parcela.num_parcela}ª parcela')
            ws.cell(row=row, column=5, value=float(parcela.valor_pago) if parcela.valor_pago is not None else 0.0)
            ws.cell(row=row, column=6, value=float(parcela.valor_parcela) if parcela.valor_parcela is not None else 0.0)
            ws.cell(row=row, column=7, value=parcela.get_situacao_display())
            row += 1
        
        return row + 2
    
    def _add_mixto_to_excel(self, ws, results, start_row):
        """Adicionar relatório misto ao Excel"""
        row = start_row
        
        # Projetos
        if results.get('projetos', {}).get('dados'):
            row = self._add_projetos_to_excel(ws, results['projetos'], row)
        
        # Contratos
        if results.get('contratos', {}).get('dados'):
            row = self._add_contratos_to_excel(ws, results['contratos'], row)
        
        return row
    
    def _add_totals_to_excel(self, ws, results, start_row):
        """Adicionar totalizadores ao Excel"""
        row = start_row + 2
        
        ws[f'A{row}'] = 'RESUMO'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        if results.get('tipo') == 'projetos':
            ws[f'A{row}'] = f'Total de Projetos: {results.get("total_registros", 0)}'
            row += 1
            ws[f'A{row}'] = f'Valor Total: R$ {results.get("valor_total", 0):.2f}'
        elif results.get('tipo') == 'contratos':
            ws[f'A{row}'] = f'Total de Contratos: {results.get("total_registros", 0)}'
            row += 1
            ws[f'A{row}'] = f'Valor Total: R$ {results.get("valor_total", 0):.2f}'
        elif results.get('tipo') == 'financeiro':
            ws[f'A{row}'] = f'Total de Parcelas: {results.get("total_registros", 0)}'
            row += 1
            ws[f'A{row}'] = f'Valor Pago: R$ {results.get("valor_total_pago", 0):.2f}'
            row += 1
            ws[f'A{row}'] = f'Valor Previsto: R$ {results.get("valor_total_previsto", 0):.2f}'
        elif results.get('tipo') == 'mixto':
            projetos = results.get('projetos', {})
            contratos = results.get('contratos', {})
            ws[f'A{row}'] = f'Projetos: {projetos.get("total_registros", 0)} (R$ {projetos.get("valor_total", 0):.2f})'
            row += 1
            ws[f'A{row}'] = f'Contratos: {contratos.get("total_registros", 0)} (R$ {contratos.get("valor_total", 0):.2f})'